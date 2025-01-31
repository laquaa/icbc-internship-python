import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

#Read
loan_2023_12 = pd.read_excel('202312反传数据.xlsx', sheet_name='贷款 ')
loan_2024_06_xls = pd.read_excel('202406返传数据.xls', sheet_name='贷款')
loan_2024_06_xls.to_excel('202406返传数据.xlsx', engine='openpyxl', index=False)
loan_2024_06 = pd.read_excel('202406返传数据.xlsx', sheet_name='Sheet1')

#Select
loan_2023_12 = loan_2023_12[['支行考核口径部门', '分行考核口径部门', '贷款余额', '币种', '小企业']]
loan_2024_06 = loan_2024_06[['支行考核口径部门', '分行考核口径部门', '贷款余额', '币种', '小企业']]

loan_2023_12 = loan_2023_12[loan_2023_12['小企业'].isna()]
loan_2024_06 = loan_2024_06[loan_2024_06['小企业'].isna()]

loan_2023_12 = loan_2023_12[loan_2023_12['币种'] == '人民币']
loan_2024_06 = loan_2024_06[loan_2024_06['币种'] == '人民币']

#department_0
department_0_2023_12 = loan_2023_12.groupby(['支行考核口径部门'], as_index=False).agg({'贷款余额': 'sum'})
department_0_2024_06 = loan_2024_06.groupby(['支行考核口径部门'], as_index=False).agg({'贷款余额': 'sum'})

#department_1
department_1_2023_12 = loan_2023_12.groupby(['分行考核口径部门'], as_index=False).agg({'贷款余额': 'sum'})
department_1_2024_06 = loan_2024_06.groupby(['分行考核口径部门'], as_index=False).agg({'贷款余额': 'sum'})

#merge
department_0_2023_12_final = department_0_2023_12.rename(columns={'贷款余额': '12月贷款余额'})
department_0_2024_06_final = department_0_2024_06.rename(columns={'贷款余额': '6月贷款余额'})
merged_department_0 = pd.merge(department_0_2023_12_final, department_0_2024_06_final, on='支行考核口径部门', how='outer')

department_1_2023_12_final = department_1_2023_12.rename(columns={'贷款余额': '12月贷款余额'})
department_1_2024_06_final = department_1_2024_06.rename(columns={'贷款余额': '6月贷款余额'})
merged_department_1 = pd.merge(department_1_2023_12_final, department_1_2024_06_final, on='分行考核口径部门', how='outer')

merged_department_0 = merged_department_0.fillna(0)
merged_department_1 = merged_department_1.fillna(0)

merged_department_0['贷款余额较年初涨跌幅'] = merged_department_0['6月贷款余额'] - merged_department_0['12月贷款余额']
merged_department_1['贷款余额较年初涨跌幅'] = merged_department_1['6月贷款余额'] - merged_department_1['12月贷款余额']

sum_department_0_2023_12 = merged_department_0['12月贷款余额'].sum()
sum_department_0_2024_06 = merged_department_0['6月贷款余额'].sum()
sum_department_0_increase = merged_department_0['贷款余额较年初涨跌幅'].sum()
new_row_department_0 = pd.DataFrame([['总计', sum_department_0_2023_12, sum_department_0_2024_06, sum_department_0_increase]], columns=merged_department_0.columns)
merged_department_0 = pd.concat([merged_department_0, new_row_department_0], ignore_index=True)

sum_department_1_2023_12 = merged_department_1['12月贷款余额'].sum()
sum_department_1_2024_06 = merged_department_1['6月贷款余额'].sum()
sum_department_1_increase = merged_department_1['贷款余额较年初涨跌幅'].sum()
new_row_department_1 = pd.DataFrame([['总计', sum_department_1_2023_12, sum_department_1_2024_06, sum_department_1_increase]], columns=merged_department_1.columns)
merged_department_1 = pd.concat([merged_department_1, new_row_department_1], ignore_index=True)

merged_df = pd.DataFrame()
merged_df = pd.concat([merged_df, merged_department_0])
empty_row = pd.DataFrame([['']*len(merged_department_0.columns)], columns=merged_department_0.columns)
merged_df = pd.concat([merged_df, empty_row], ignore_index=True)
column_names = pd.DataFrame([merged_department_1.columns.tolist()], columns=merged_department_0.columns)
merged_df = pd.concat([merged_df, column_names], ignore_index=True)
merged_department_1.columns = merged_department_0.columns
all_departments = pd.concat([merged_df, merged_department_1], ignore_index=True)

#title
new_row_0 = pd.DataFrame([['非小企业人民币贷款按不同口径核算', '', '', '']])
new_row_1 = pd.DataFrame([['统计时间节点', '20240630', '', '']])
new_row_2 = pd.DataFrame([['单位', '人民币万元', '', '']])
title = pd.concat([new_row_0, new_row_1, new_row_2], ignore_index=True)

merged_df = pd.DataFrame()
merged_df = pd.concat([merged_df, title])
empty_row = pd.DataFrame([['']*len(title.columns)], columns=title.columns)
merged_df = pd.concat([merged_df, empty_row], ignore_index=True)
column_names = pd.DataFrame([all_departments.columns.tolist()], columns=title.columns)
merged_df = pd.concat([merged_df, column_names], ignore_index=True)
all_departments.columns = title.columns
all_departments = pd.concat([merged_df, all_departments], ignore_index=True)

#save
all_departments.to_excel('非小企业人民币贷款统计（部门）.xlsx', index=False, header=False)

#edit
wb = openpyxl.load_workbook('非小企业人民币贷款统计（部门）.xlsx')
ws = wb.active

merge_cells_coords = [
    ('A1', 'B1')
]
for start_cell, end_cell in merge_cells_coords:
    ws.merge_cells(f'{start_cell}:{end_cell}')
    cell = ws[start_cell]
    cell.alignment = Alignment(horizontal='center', vertical='center')

bold_font = Font(bold=True)

rows_to_format = [1, 2, 3]
columns_to_format = ['中国###总集', 'B']

for row in rows_to_format:
    for col in columns_to_format:
        cell = ws[f'{col}{row}']
        cell.font = bold_font
for col in ws.iter_cols(min_row=5, max_row=5):
    for cell in col:
        cell.font = bold_font
for col in ws.iter_cols(min_row=12, max_row=12):
    for cell in col:
        cell.font = bold_font

black_side = Side(border_style="thin", color="000000")
black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)

for row in rows_to_format:
    for col in columns_to_format:
        cell = ws[f'{col}{row}']
        cell.border = black_border
for row in ws.iter_rows(min_row=5, max_row=5):
    for cell in row:
        cell.border = black_border
for row in ws.iter_rows(min_row=12, max_row=12):
    for cell in row:
        cell.border = black_border

wb.save('非小企业人民币贷款统计（部门）.xlsx')